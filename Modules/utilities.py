
import pickle 
import os 
from openpyxl import load_workbook

from PyQt5.QtWidgets import (
    QFileDialog
)


ERROR_CODES = { 
    0: 'test'
}

def openFile(): 
    try: 
        fileName, _ = QFileDialog.getOpenFileName(None, 'Open File', '',)
        return fileName
    except: 
        return None; 
    
def getFileLocation():
    dlg = QFileDialog().getExistingDirectory()
    print(dlg)
    return dlg


class LocalPreferences: 
    def __init__(self, path='data.pkl'): 
        self.path = path
        self.load()
    
    def load(self):
        try:
            with open(self.path, 'rb') as file:
                self.preferences = pickle.load(file)
        except (FileNotFoundError, EOFError):
            print('Error: could not load prefernces')
            self.preferences = {}
            
    def values(self): 
        return self.preferences
    
    def update(self,name, value): 
        self.preferences[name] = value
        self.save()
        
    def get(self, value): 
        return self.preferences[value]
    
    def remove(self, value):
        del self.preferences[value]
        
    def save(self): 
        with open(self.path, 'wb') as file:
            pickle.dump(self.preferences, file)

def saveLocation(data): 

    fileName = 'data.pkl'
    #locations = loadLocations()
    #locations[locationName] = location; 
    
    with open(fileName, 'wb') as file: 
        pickle.dump(data, file)
        
def loadLocations(): 
    
    fileName = 'data.pkl'

    with open(fileName, 'rb') as file: 
        load_data = pickle.load(file)
        return load_data
    
    #return load_data; 


def fileExtenCheck(filePath): 
    if(os.path.isfile(filePath)): 
        fileExtension = os.path.splitext(filePath)[1].lower()
        
        if fileExtension in ['.xlsx', '.csv']: 
            return True; 
        else: 
            return False; 
        
    return False; 


def updateLOQ(): 
    picklePath = 'LOQ_DATA.pkl'

    filePath = openFile()
    
    if filePath: 
        wb = load_workbook(filename=filePath)
        ws = wb.active
        
        columns_to_read = [1,2]  # Columns A and C (0-based index)
        startingRow = 2

        LOQ_DATA = {}
        

        # Read data from the specified columns
        print('**LOADING LOQ FORM DATA')
        for row in ws.iter_rows(min_row=startingRow, values_only=True, min_col=2 , max_col=5):
            print(row)
            compoundName = row[0]
            
            budValue = row[1]
            oilValue = row[2]
            paperValue = row[3]

            LOQ_DATA[compoundName] = [
                budValue, 
                oilValue,
                paperValue
            ]
        
        #print(LOQ_DATA)
        with open(picklePath, 'wb') as file: 
            pickle.dump(LOQ_DATA, file) 
        
        
        wb.close()
    
def loadLOQ(): 
    picklePath = 'LOQ_DATA.pkl'
    
    with open(picklePath, 'rb') as file: 
        loaded_data = pickle.load(file)
    
    return loaded_data
    
    #try block 
    
    