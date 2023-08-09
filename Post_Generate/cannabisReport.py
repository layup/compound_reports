import pickle 
import os 
import math; 

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, borders, Border, Side
from openpyxl.cell.rich_text import TextBlock, CellRichText 
from openpyxl.worksheet.page import PageMargins
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

from Modules.utilities import * 
from Post_Generate.excel import * 


THC_NAMES = {
    'Ibuprofen': '',
    'CBDV': 'Cannabidivarin (CBDV)',
    'CBDVA': 'Cannabidivarin Acid',
    'THCV': 'Δ⁹ THCV',
    'CBGVA': '', #unsure what this one represents 
    'CBD': 'Cannabidiol (CBD)',
    'CBG': 'Cannabigerol (CBG)',
    'CBGA': 'Cannabigerol Acid',
    'CBDA': 'Cannabidol Acid',
    'CBN': 'Cannabinol (CBN)',
    'THCVA': 'Δ⁹ THCV Acid',
    'd9-THC': 'Δ⁹ THC',
    'd8-THC': 'Δ⁸ THC',
    'CBL': 'Cannabicyclol (CBL)',
    'CBC': 'Cannabichromene (CBC)',
    'CBCA': 'Cannabichromene Acid',
    'CBNA': 'Cannabional Acid',
    'THCA': 'Δ⁹ THC Acid',
    'CBLA': 'Cannabicyclol Acid',
    '*': 'Total THC*', 
    '**': 'Total CBD**', 

}

basicTestsFormatting = [
    'd9-THC', 
    'THCA', 
    '*',
    'd8-THC', 
    'CBD', 
    'CBDA', 
    '**', 
    'CBN', 
    'CBNA' 
]

dexlueTestsFormatting = [
    'd9-THC', 
    'THCA', 
    '*',
    'd8-THC', 
    'CBC', 
    'CBCA', 
    'CBD', 
    'CBDA', 
    '**',
    'CBG', 
    'CBGA', 
    'CBL', 
    'CBLA',
    'CBDV',
    'CBDVA',
    'THCV', 
    'THCVA', 
    'CBN', 
    'CBNA'
]

basicHeaders = [
    THC_NAMES['d9-THC'],
    THC_NAMES['THCA'],
    'Total THC*',
    THC_NAMES['d8-THC'],
    THC_NAMES['CBD'], 
    THC_NAMES['CBDA'], 
    'Total CBD**', 
    THC_NAMES['CBN'],
    THC_NAMES['CBNA']         
]

deluxeHeaders = [THC_NAMES.get(name, name) for name in dexlueTestsFormatting]


unitTypeMoisture = 0; 
unitTypeDensity = 1; 
unitTypeMass = 2; 
unitTypePercentage = 3; 

basicReport = 0 
deluxeReport = 1; 

def reportPaths(jobNums, samples, excelFileName):
   
    results = loadLocations(); 
    path = results['output'];  

    outputPaths = {}
    
    textFileName = 'source.txt'
    textContent = f'Source File {excelFileName}' 

    for jobNum, value in samples.items(): 
        
        folderPath = os.path.join(path, jobNum)
        if not os.path.exists(folderPath):
            os.makedirs(folderPath)
            textPath = os.path.join(folderPath, textFileName)
            try: 
                with open(textPath, "w") as file:
                    file.write(textContent)
            except: 
                pass; 

            

        multiFileName = 'W' + str(jobNum) + "_thc.xlsx"
        multiFilePath =  os.path.join(folderPath, multiFileName)
            
        singlePath = {}
        
        for pathType, value2 in value.items(): 

            if(pathType == 'single'):
                for item in value2: 
                    singleFileName = 'W' + str(item) + "_thc.xlsx"
                    singleFilePath =  os.path.join(folderPath, singleFileName) 
                    
                    singlePath[item] = singleFilePath
                
        outputPaths[jobNum] = {
            'multi': multiFilePath, 
            'single': singlePath
        }


    return outputPaths 

#TODO: Recovery 

def generateThcReport(jobNums, clientInfo, sampleInfo, sampleData, recoveryValues, excelFileName):
    print('****Generating THC Options')
    print('***Job Numbers: ', jobNums)
    
    for sampleNum, sampleValue in sampleInfo.items():
        print('{}:{}'.format(sampleNum, sampleValue))
    
    sampleType = {}
    sampleNames = {}

    reportType = {} 
        
    unitType = {}
    
    #TODO: insert the special test cases 
    unitValue = {}
    unitMassType = {}
    showExtraRow = {}
    
    for job in jobNums: 
        single = []
        multi = []

        for key, value in sampleInfo.items(): 
   
            if job in key:
                if value[3] == 'Multi': 
                    multi.append(key)
                else: 
                    single.append(key)
                
                if value[2] == 'Basic Report': 
                    reportType[job] = basicReport
                else: 
                     reportType[job] = deluxeReport
                     
                if '(Moisture)' in value[1]: 
                    unitType[job] = unitTypeMoisture
        
                if '(Density)' in value[1]: 
                    unitType[job] = unitTypeDensity
                    
                if '(Unit Mass)' in value[1]: 
                    unitType[job] = unitTypeMass
                    
                if 'Percent Only' in value[1]: 
                    unitType[job] = unitTypePercentage 
                
                sampleNames[key] = value[6]; 
                
                if(value[4] != ''): 
                    unitValue[key] = value[4] 
                    showExtraRow[job] = True;
                else: 
                    if(job not in showExtraRow): 
                        showExtraRow[job] = False; 
                
                if(value[5] != ''): 
                    unitMassType[job] = value[5]; 
                               
        sampleType[job]  = {
            'single':single, 
            'multi': multi
        }
    
    print('**Mult or Sample')
    for key, value in sampleType.items(): 
        print(key, value)
    
    print('**Report Type')
    for key, value in reportType.items(): 
        print(key, value)

    print('**Unit Type') 
    for key, value in unitType.items(): 
        print(key, value)
    
    print('**Sample Names')
    for key, value in sampleNames.items(): 
        print(key, value)

    print('**Show Extra Row')  
    for key, value in showExtraRow.items(): 
        print(key, value)

    print('**Special Values')
    for key, value in unitValue.items(): 
        print(key, value)
    
    print('** Unit Mass Type Values')
    for key, value in unitMassType.items():
        print(key, value)

    outputPaths = reportPaths(jobNums, sampleType, excelFileName)
    
    print('***OuputPaths')
    for jobNum, value in outputPaths.items(): 
        print(jobNum, value)
    
    #FIXME: single's are not working at the moment 
    for jobNum, report in sampleType.items(): 
        print('****Current Job: ', jobNum)
        print('Mult/Single: ', report)
        print('report type:', reportType[jobNum])
        print('Unit Type: ', unitType[jobNum]) 
        
        for currentReportType, samples in report.items(): 
            print('Report: ', currentReportType, '| Samples: ', samples )
            if(currentReportType == 'single'): 
                for currentSample in samples: 
                    singleOutputPath = outputPaths[jobNum]['single'][currentSample]
                    
                    if(reportType[jobNum] == basicReport): 
                        temp = [currentSample]
                        generateBasicReport(jobNum, singleOutputPath, clientInfo[jobNum], temp, sampleNames, sampleData, recoveryValues, unitType[jobNum], showExtraRow[jobNum], unitValue, unitMassType )
                    else: 
                        temp = [currentSample] 
                        generateDexluxeReport(jobNum, singleOutputPath, clientInfo[jobNum], temp, sampleNames, sampleData, recoveryValues, unitType[jobNum], showExtraRow[jobNum], unitValue, unitMassType )
                    print('----------------------------') 
            else:
                if(len(samples) != 0 ): 
                    mulitOutputPath = outputPaths[jobNum]['multi'] 
                    
                    if(reportType[jobNum] == basicReport): 
                        generateBasicReport(jobNum, mulitOutputPath, clientInfo[jobNum], samples, sampleNames, sampleData, recoveryValues, unitType[jobNum], showExtraRow[jobNum], unitValue, unitMassType )
                    else: 
                        generateDexluxeReport(jobNum, mulitOutputPath, clientInfo[jobNum], samples, sampleNames, sampleData, recoveryValues, unitType[jobNum], showExtraRow[jobNum], unitValue, unitMassType ) 
                    print('----------------------------') 
                              

  
def generateBasicReport(jobNum, outputPath, clientInfo, samples, sampleNames, sampleData, recoveryValues, unitType, showExtraRow, unitValue, unitMassType=None): 
    print('****Creating Basic THC Reports: ', jobNum)
    print('**Samples: ', samples)
    print('**Sample Names: ', sampleNames)
    print('*Output Path: ', outputPath)
    print('*unitType', unitType)
    #print(recoveryValues)
    print('*Sample Data')
    print(sampleData)
    print('*Unit Value')
    print(unitValue)
    print('*Unit Mass Type')
    print(unitMassType)
    
    wb = Workbook()
    ws = wb.active 
    
    pageSetup(ws);    
    headerTitle = 'Cannabis Basic Report'
    createFooters(ws, headerTitle, jobNum); 
    
    ws = insertClientInfo(ws, clientInfo, 'D')
    
    for column in range(1, 9): 
        columnLetter = get_column_letter(column)
        
        if(column == 1):
            ws.column_dimensions[columnLetter].width = 20 
        else: 
            ws.column_dimensions[columnLetter].width = 11
        
    ws.print_title_rows = '1:8' # the first two rows
    
    #Basic Page information 
    pageSize = 56; 
    maxCol = 8
    pageLocation = 9; 
    currentPage = 1; 
    
    totalSamples = len(samples)
    totalPages = determineTotalPages(totalSamples, unitType) 
    
    print('Total Samples: ', totalSamples, ' Total Pages: ', totalPages); 
     
    #format the given rows   
    formatRows(ws, pageSize, totalPages, maxCol)  

    #generate the basic names 
    currentSection = 0;
    sampleNumber  =1; 
    sectionJobs = calculateSections(samples, totalSamples, unitType)
    sampleNames = generateSampleSectionNames(samples, sampleNames, unitType) 
    
    totalSections = len(sectionJobs)
    
    print("**Section Breakdown")
    for section in sectionJobs: 
        print(section)

    print("**Section Sample Names")
    for sampleName in sampleNames: 
        print(sampleName)
         
    
    for i, section in enumerate(sectionJobs, 1):  
        pageLocation = insertSampleName(ws, pageLocation, sampleNames[currentSection],8)
        pageLocation, sampleNumber = insertBasicTableHeader(ws, pageLocation, sectionJobs, currentSection, sampleNumber, unitType)
        insertBasicTable(ws, pageLocation, recoveryValues, showExtraRow)
    
        pageLocation = insertSampleValues(ws, pageLocation, sectionJobs, currentSection, sampleData, unitType, unitValue)
        pageLocation = insertExtraRowValues(ws, jobNum, pageLocation, sectionJobs, currentSection, showExtraRow, unitType, unitValue, unitMassType)
      
        if(i == totalSections): 
            pageLocation +=1; 
            insertThcComment(ws, pageLocation) 
            break;  
        
        currentSection += 1;  
            
        if(i+1 == totalSections): 
            if(i % 2 == 0):
                pageLocation+=1; 
                insertNextSectionComment(ws, pageLocation)    
                currentPage+=1; 
                pageLocation = ((currentPage-1) * pageSize) - (8 * (currentPage-2)) + 1;  
                   
        else: 
            if(i % 2 == 0): 
                pageLocation+=1; 
                insertNextSectionComment(ws, pageLocation)    
                currentPage+=1; 
                pageLocation = ((currentPage-1) * pageSize) - (8 * (currentPage-2)) + 1;  

            

    wb.save(outputPath) 
    
    
def generateDexluxeReport(jobNum, outputPath, clientInfo, samples, sampleNames, sampleData, recoveryValues, unitType, showExtraRow, unitValue, unitMassType=None): 
    print('****Creating Deluxe THC Reports: ', jobNum)
    print('**Samples: ', samples)
    print('**Sample Names: ', sampleNames)
    print('*Output Path: ', outputPath)
    print('*unitType', unitType)
    #print(recoveryValues)
    print('*Sample Data')
    print(sampleData)
    print('*Unit Value')
    print(unitValue)
    print('*Unit Mass Type')
    print(unitMassType) 
    
    wb = Workbook()
    ws = wb.active 
    
    pageSetup(ws);    
    headerTitle = 'Cannabis Deluxe Report'
    createFooters(ws, headerTitle, jobNum); 
    
    ws = insertClientInfo(ws, clientInfo, 'D')
    
    for column in range(1, 9): 
        columnLetter = get_column_letter(column)
        
        if(column == 1):
            ws.column_dimensions[columnLetter].width = 20 
        else: 
            ws.column_dimensions[columnLetter].width = 11
        
    ws.print_title_rows = '1:8' # the first two rows
    
    #Basic Page information 
    pageSize = 56; 
    maxCol = 8
    pageLocation = 9; 
    currentPage = 1; 
    
    totalSamples = len(samples)
    totalPages = determineTotalPages2(totalSamples, unitType) 
    
    print('Total Samples: ', totalSamples, ' Total Pages: ', totalPages); 
     
    #format the given rows   
    formatRows(ws, pageSize, totalPages, maxCol)  

    #generate the basic names 
    currentSection = 0;
    sampleNumber  =1; 
    sectionJobs = calculateSections(samples, totalSamples, unitType)
    sampleNames = generateSampleSectionNames(samples, sampleNames, unitType) 
    
    totalSections = len(sectionJobs)
    
    print("**Section Breakdown")
    for section in sectionJobs: 
        print(section)

    print("**Section Sample Names")
    for sampleName in sampleNames: 
        print(sampleName)

    #for i, section in enumerate(sectionJobs, 1):  
    for i, section in enumerate(sectionJobs, 1):   
        pageLocation = insertSampleName(ws, pageLocation, sampleNames[currentSection], 8)
        pageLocation, sampleNumber = insertBasicTableHeader(ws, pageLocation, sectionJobs, currentSection, sampleNumber, unitType)
        insertDeluxeTable(ws, pageLocation, recoveryValues, showExtraRow)
    
        pageLocation = insertSampleValues2(ws, pageLocation, sectionJobs, currentSection, sampleData, unitType, unitValue)
        pageLocation = insertExtraRowValues(ws, jobNum, pageLocation, sectionJobs, currentSection, showExtraRow, unitType, unitValue, unitMassType)


        if(i == totalSections): 
            pageLocation +=1; 
            insertThcComment(ws, pageLocation) 
        else: 
            currentSection += 1;
             
            pageLocation +=1; 
            insertNextSectionComment(ws, pageLocation);
            currentPage +=1; 
            pageLocation = ((currentPage-1) * pageSize) - (8 * (currentPage-2)) + 1;  
     
    
    
    wb.save(outputPath); 

def insertExtraRowValues(ws, jobNum, pageLocation, sectionJobs, currentSection, showExtraRow, unitType, unitValue, unitMassType): 
    if(unitType in [0,1,2]): 
        if(showExtraRow): 
            if(unitType in [0,1]):
                
                unitNameOptions = [
                    'Moisture (%)', 
                    'Density (g/mL)' 
                ]
                
                rowName = ws.cell(row=pageLocation, column=1)
                rowName.alignment = Alignment(horizontal='left', indent=1) 
                rowName.value = unitNameOptions[unitType]
                rowName.font = Font(name="Times New Roman", size=9, bold=True) 
                
                for index, sample in enumerate(sectionJobs[currentSection]): 
                    if(index == 0): 
                        currentColumn = 2; 
                    else: 
                        currentColumn = 4;   

                    currentSampleRow = ws.cell(row=pageLocation, column=currentColumn)
                    currentSampleRow.alignment = Alignment(horizontal='left', indent=1)  
                    currentSampleRow.font = Font(name="Times New Roman", size=9, bold=True) 
                    
                    try: 
                        print('UnitValue: ', unitValue[sample])
                        result = formatLeadingValues(float(unitValue[sample]))
                        currentSampleRow.value = result; 
                    except: 
                        print('There was an error inserting ', unitNameOptions[unitType], ' for ', sample)
                    
    
            if(unitType == 2): 
                rowName = ws.cell(row=pageLocation, column=1)
                rowName.alignment = Alignment(horizontal='left', indent=1) 
                rowName.font = Font(name="Times New Roman", size=9, bold=True) 

                if(str(jobNum) in unitMassType): 
                    rowName.value = f'Unit Mass ({unitMassType[str(jobNum)]})'
                else: 
                    rowName.value = 'Unit Mass (unit/mass)' 
                
                for index, sample in enumerate(sectionJobs[currentSection]): 
                    if(index == 0): 
                        currentColumn = 2; 
                    else: 
                        currentColumn = 4;   

                    currentSampleRow = ws.cell(row=pageLocation, column=currentColumn)
                    currentSampleRow.alignment = Alignment(horizontal='left', indent=1)  
                    currentSampleRow.font = Font(name="Times New Roman", size=9, bold=True) 
                    
                    try: 
                        result = formatLeadingValues(float(unitValue[sample]))
                        currentSampleRow.value = result; 
                    
                    except: 
                        print('There was an error inserting Unit Mass for ', sample)
                    
            pageLocation +=2; 
        else: 
            pageLocation += 1; 
    else: 
        pageLocation +=1; 
            
    return pageLocation; 


def combineBorder(cell, newBorder): 

    if(cell.border): 
        left_border_style = cell.border.left.style if cell.border.left else None
        right_border_style = cell.border.right.style if cell.border.right else None
        top_border_style = cell.border.top.style if cell.border.top else None
        bottom_border_style = cell.border.bottom.style if cell.border.bottom else None
        
        combined_border = Border(
            left=cell.border.left if left_border_style else newBorder.left,
            right=cell.border.right if right_border_style else newBorder.right,
            top=cell.border.top if top_border_style else newBorder.top,
            bottom=cell.border.bottom if bottom_border_style else newBorder.bottom
        )
        return combined_border
     
    return  newBorder; 
        

def determineTotalPages(totalSamples, unitType): 
    if(unitType in [0,1,2]):
        samplesPerPage = 4
        return math.ceil(totalSamples/samplesPerPage)
    else: 
        samplesPerPage = 8
        return math.ceil(totalSamples/samplesPerPage)
    
def determineTotalPages2(totalSamples, unitType): 
    if(unitType in [0,1,2]):
        samplesPerPage = 2
        return math.ceil(totalSamples/samplesPerPage)
    else: 
        samplesPerPage = 4
        return math.ceil(totalSamples/samplesPerPage)
    
def calculateSections(samples, totalSamples, unitType): 
    sectionJobs = []; 
    
    if(unitType in [0,1,2]): 
        headerLimit = 2 
    else: 
        headerLimit = 4 
    
    if(totalSamples <= headerLimit): 
        sectionJobs.append(samples) 
    else: 
        tempSection = []
        for i, sample in enumerate(samples, start=1): 
            
            tempSection.append(sample)
            if(i % headerLimit == 0): 
                sectionJobs.append(tempSection)
                tempSection = []
                
        if(len(tempSection) != 0): 
            sectionJobs.append(tempSection);
            
    return sectionJobs

def insertBasicTable(ws, pageLocation, recoveryValues, showExtraRow ): 
    
    for i, headerName in enumerate(basicHeaders): 
        testNameColumn = ws.cell(row=pageLocation + i, column=1)
        testNameColumn.value = headerName
        testNameColumn.alignment = Alignment(indent=1)   
        testNameColumn.border = Border(right=thinBorder)
        
        if("*" in headerName): 
            testNameColumn.font = Font(name="Times New Roman", size=9, bold=True) 

        blankColumn = ws.cell(row=pageLocation + i , column=6)
        blankColumn.alignment = Alignment(indent=1)   
        blankColumn.border = Border(left=thinBorder)

        #TODO: check if 0.001 is the exact thing
        soColumn = ws.cell(row=pageLocation + i, column=8) 
        soColumn.alignment = Alignment(horizontal='left',indent=1)
        
        if(i not in [2,6]): 
            blankColumn.value = 'ND'
            soColumn.value = 0.001

        if(i in[2,3,6]): 
            for col in range(1,9): 
                temp = ws.cell(row=pageLocation + i, column=col)
                newBorder = Border(top=thinBorder, bottom=thinBorder)
                temp.border = combineBorder(temp, newBorder)

        if(i == 8): 
            for col in range(1,9): 
                temp = ws.cell(row=pageLocation + i, column=col)
                newBorder = Border(bottom=thinBorder)
                temp.border = combineBorder(temp, newBorder)
            
    print('**Recovery Values')
    for index, test in enumerate(basicTestsFormatting): 
        recoveryColumn = ws.cell(row=pageLocation + index, column=7)
        recoveryColumn.alignment = Alignment(horizontal='left', indent=1)

        if(test in recoveryValues.keys()): 
    
            recoveryVal = float(recoveryValues[test])
            recoveryColumn.value = significantFiguresConvert(recoveryVal) 
            
            number_format = get_format_for_value(recoveryVal)
            #recoveryColumn.number_format = number_format  

            print(f'Recovery [{test}]: {number_format}, {recoveryVal}')
    
    #insert additonal column 
    #TODO: if unitType = 4 do not include this 
    if(showExtraRow): 
        for col in range(1,9): 
            blankRow = ws.cell(row = pageLocation + 9, column=col); 
            blankRow.alignment = Alignment(horizontal='left', indent=1)
            blankRow.border = Border(bottom=thinBorder)
            
            if(col in [1,3,5]): 
                blankRow.border = Border(bottom=thinBorder, right=thinBorder) 
           
def insertDeluxeTable(ws, pageLocation, recoveryValues, showExtraRow): 
    
    for i, headerName in enumerate(deluxeHeaders): 
        testNameColumn = ws.cell(row=pageLocation + i, column=1)
        testNameColumn.value = headerName
        testNameColumn.alignment = Alignment(indent=1)   
        testNameColumn.border = Border(right=thinBorder)

        if("*" in headerName): 
            testNameColumn.font = Font(name="Times New Roman", size=9, bold=True) 

        blankColumn = ws.cell(row=pageLocation + i , column=6)
        blankColumn.alignment = Alignment(indent=1)   
        blankColumn.border = Border(left=thinBorder)

        #TODO: check if 0.001 is the exact thing
        soColumn = ws.cell(row=pageLocation + i, column=8) 
        soColumn.alignment = Alignment(horizontal='left',indent=1)

        if(i not in [2,8]): 
            blankColumn.value = 'ND'
            soColumn.value = 0.001

        if(i in[2,3,8]): 
            for col in range(1,9): 
                temp = ws.cell(row=pageLocation + i, column=col)
                newBorder = Border(top=thinBorder, bottom=thinBorder)
                temp.border = combineBorder(temp, newBorder)

        if(i in [5,10,12,14,16,18]): 
            for col in range(1,9): 
                temp = ws.cell(row=pageLocation + i, column=col)
                newBorder = Border(bottom=thinBorder)
                temp.border = combineBorder(temp, newBorder)
            
    print('**Recovery Values')
    for index, test in enumerate(dexlueTestsFormatting): 
        recoveryColumn = ws.cell(row=pageLocation + index, column=7)
        recoveryColumn.alignment = Alignment(horizontal='left', indent=1)

        if(test in recoveryValues.keys()): 
    
            recoveryVal = float(recoveryValues[test])
            recoveryColumn.value = significantFiguresConvert(recoveryVal) 
            
            number_format = get_format_for_value(recoveryVal)
            #recoveryColumn.number_format = number_format  

            print(f'Recovery [{test}]: {number_format}, {recoveryVal}')
    
    #insert additonal column 
    #TODO: if unitType = 4 do not include this 
    if(showExtraRow): 
        for col in range(1,9): 
            blankRow = ws.cell(row = pageLocation + len(deluxeHeaders), column=col); 
            blankRow.alignment = Alignment(horizontal='left', indent=1)
            blankRow.border = Border(bottom=thinBorder)
            
            if(col in [1,3,5]): 
                blankRow.border = Border(bottom=thinBorder, right=thinBorder) 

        
    
            
def insertSampleValues(ws, pageLocation, sectionJobs, currentSection, sampleData, unitType, unitValue=None ): 
    unitTypeOptions = [
        ['(mg/g)', '(%)'], 
        ['(mg/mL)', '(%)'], 
        ['(mg/g)', '(mg/unit)'], 
        ['(%)', '(%)']
    ]
   
    for index, sample in enumerate(sectionJobs[currentSection]): 
        percentageMulti = 0.877
        currentSample = sampleData[sample]
        print('Current Sample: ', sample)
        #print('***Current Sample: ', currentSample) 
        if(unitType in [0,1]): 
            if(index == 0): 
                currentColumn = 2; 
            else: 
                currentColumn = 4;  

            for i, test in enumerate(basicTestsFormatting): 
                valueColumn = ws.cell(row=pageLocation+i, column=currentColumn)
                valueColumn.alignment = Alignment(horizontal='left', indent=1)


                percentageColumn = ws.cell(row=pageLocation+i, column=currentColumn+1)
                percentageColumn.alignment = Alignment(horizontal='left', indent=1)
                
                newBorder = Border(left=thinBorder, right=thinBorder)
                percentageColumn.border = combineBorder(percentageColumn, newBorder) 

                if("*" in test): 
                    valueColumn.font = Font(name="Times New Roman", size=9, bold=True) 
                    percentageColumn.font =Font(name="Times New Roman", size=9, bold=True)  
                
                if(test in currentSample.keys()): 
                    testVal = currentSample[test]

                    print('Tests: %s | Value: %s ' % (test, testVal) )

                    if(testVal == 0.0): 
                        valueColumn.value = 'ND'
                        percentageColumn.value = 'ND'
                    else: 
                        processValues(testVal, unitType, valueColumn, percentageColumn)
                    
                else: 
                    if(test == '*'): 
                        thc = currentSample['d9-THC']
                        thcv = currentSample['THCA'] * percentageMulti 
                        total_thc = thc + thcv 
                    
                        print('Tests: %s | Value: %s ' % (test, total_thc))
                        processValues(total_thc, unitType, valueColumn, percentageColumn)
                        
                    if(test == '**'): 
                        cbd = currentSample['CBD']
                        cbda = currentSample['CBDA'] * percentageMulti 
                        total_cbd = cbd + cbda 

                        print('Tests: %s | Value: %s ' % (test, total_cbd))
                        processValues(total_cbd, unitType, valueColumn, percentageColumn)
                        
        if(unitType == 2): 
            if(index == 0): 
                currentColumn = 2; 
            else: 
                currentColumn = 4;   
            
            if(sample in unitValue): 
                unitMass = float(unitValue[sample])
            else: 
                unitMass = 1; 
            
            for i, test in enumerate(basicTestsFormatting): 
                valueColumn = ws.cell(row=pageLocation+i, column=currentColumn)
                valueColumn.alignment = Alignment(horizontal='left', indent=1)
                
                valueMultColumn = ws.cell(row=pageLocation+i, column=currentColumn+1)
                valueMultColumn.alignment = Alignment(horizontal='left', indent=1)

                newBorder = Border(left=thinBorder, right=thinBorder)
                valueMultColumn.border = combineBorder(valueMultColumn, newBorder) 

                if("*" in test): 
                    valueColumn.font = Font(name="Times New Roman", size=9, bold=True) 
                    valueMultColumn.font =Font(name="Times New Roman", size=9, bold=True)  

                if(test in currentSample.keys()): 
                    testVal = currentSample[test] 
                    print('Tests: %s | Value: %s ' % (test, testVal) )
                    
                    if(testVal == 0.0): 
                        valueColumn.value = 'ND'
                        valueMultColumn.value = 'ND'
                    else: 
                        processValues(testVal, unitType, valueColumn, valueMultColumn, unitMass)
                    
                else: 
                    if(test == '*'): 
                        thc = currentSample['d9-THC']
                        thcv = currentSample['THCA']  * percentageMulti 
                        total_thc = thc + thcv 
                    
                        print('Tests: %s | Value: %s ' % (test, total_thc))
                        processValues(total_thc, unitType, valueColumn, valueMultColumn, unitMass)
                        
                    if(test == '**'): 
                        cbd = currentSample['CBD']  
                        cbda = currentSample['CBDA'] * percentageMulti 
                        total_cbd = cbd + cbda 

                        print('Tests: %s | Value: %s ' % (test, total_cbd))
                        processValues(total_cbd, unitType, valueColumn, valueMultColumn, unitMass)
                

        if(unitType == 3): 
            currentColumn= 2 + index; 
            #TODO: if there isn't info the nwe can just have one, based on unit
            for i, test in enumerate(basicTestsFormatting): 
                
                valueColumn = ws.cell(row=pageLocation+i, column=currentColumn)
                valueColumn.alignment = Alignment(horizontal='left', indent=1)
                
                newBorder = Border(left=thinBorder, right=thinBorder)
                valueColumn.border = combineBorder(valueColumn, newBorder)  

                
                if("*" in test): 
                    valueColumn.font = Font(name="Times New Roman", size=9, bold=True) 

                if(test in currentSample.keys()): 
                    testVal = currentSample[test]
                    print('Tests: %s | Value: %s ' % (test, testVal) )

                    if(testVal == 0.0): 
                        valueColumn.value = 'ND' 
                    else: 
                        processValues(testVal, unitType, valueColumn)
                    
                else: 
                    if(test == '*'): 
                        thc = currentSample['d9-THC']
                        thcv = currentSample['THCA'] * percentageMulti 
                        total_thc = thc + thcv 
                    
                        print('Tests: %s | Value: %s ' % (test, total_thc))
                        processValues(total_thc, unitType, valueColumn)
                    
                    if(test == '**'): 
                        cbd = currentSample['CBD']
                        cbda = currentSample['CBDA'] * percentageMulti 
                        total_cbd = cbd + cbda 

                        print('Tests: %s | Value: %s ' % (test, total_cbd))
                        processValues(total_cbd, unitType, valueColumn)
    
        
    pageLocation += len(basicHeaders);
        
    return pageLocation 
        
def insertSampleValues2(ws, pageLocation, sectionJobs, currentSection, sampleData, unitType, unitValue=None ): 
    unitTypeOptions = [
        ['(mg/g)', '(%)'], 
        ['(mg/mL)', '(%)'], 
        ['(mg/g)', '(mg/unit)'], 
        ['(%)', '(%)']
    ]
   
    for index, sample in enumerate(sectionJobs[currentSection]): 
        percentageMulti = 0.877
        currentSample = sampleData[sample]
        print('Current Sample: ', sample)
        #print('***Current Sample: ', currentSample) 
        if(unitType in [0,1]): 
            if(index == 0): 
                currentColumn = 2; 
            else: 
                currentColumn = 4;  

            for i, test in enumerate(dexlueTestsFormatting): 
                valueColumn = ws.cell(row=pageLocation+i, column=currentColumn)
                valueColumn.alignment = Alignment(horizontal='left', indent=1)

                percentageColumn = ws.cell(row=pageLocation+i, column=currentColumn+1)
                percentageColumn.alignment = Alignment(horizontal='left', indent=1)
                
                newBorder = Border(left=thinBorder, right=thinBorder)
                percentageColumn.border = combineBorder(percentageColumn, newBorder) 

                if("*" in test): 
                    valueColumn.font = Font(name="Times New Roman", size=9, bold=True) 
                    percentageColumn.font =Font(name="Times New Roman", size=9, bold=True)  
                
                if(test in currentSample.keys()): 
                    testVal = currentSample[test]

                    print('Tests: %s | Value: %s ' % (test, testVal) )

                    if(testVal == 0.0): 
                        valueColumn.value = 'ND'
                        percentageColumn.value = 'ND'
                    else: 
                        processValues(testVal, unitType, valueColumn, percentageColumn)
                    
                else: 
                    if(test == '*'): 
                        thc = currentSample['d9-THC']
                        thcv = currentSample['THCA'] * percentageMulti 
                        total_thc = thc + thcv 
                    
                        print('Tests: %s | Value: %s ' % (test, total_thc))
                        processValues(total_thc, unitType, valueColumn, percentageColumn)
                        
                    if(test == '**'): 
                        cbd = currentSample['CBD']
                        cbda = currentSample['CBDA'] * percentageMulti 
                        total_cbd = cbd + cbda 

                        print('Tests: %s | Value: %s ' % (test, total_cbd))
                        processValues(total_cbd, unitType, valueColumn, percentageColumn)
                        
        if(unitType == 2): 
            if(index == 0): 
                currentColumn = 2; 
            else: 
                currentColumn = 4;   
            
            if(sample in unitValue): 
                unitMass = float(unitValue[sample])
            else: 
                unitMass = 1; 
            
            for i, test in enumerate(dexlueTestsFormatting): 
                valueColumn = ws.cell(row=pageLocation+i, column=currentColumn)
                valueColumn.alignment = Alignment(horizontal='left', indent=1)
                
                valueMultColumn = ws.cell(row=pageLocation+i, column=currentColumn+1)
                valueMultColumn.alignment = Alignment(horizontal='left', indent=1)

                newBorder = Border(left=thinBorder, right=thinBorder)
                valueMultColumn.border = combineBorder(valueMultColumn, newBorder) 
                
                if("*" in test): 
                    valueColumn.font = Font(name="Times New Roman", size=9, bold=True) 
                    valueMultColumn.font =Font(name="Times New Roman", size=9, bold=True)  

                if(test in currentSample.keys()): 
                    testVal = currentSample[test] 
                    print('Tests: %s | Value: %s ' % (test, testVal) )
                    
                    if(testVal == 0.0): 
                        valueColumn.value = 'ND'
                        valueMultColumn.value = 'ND'
                    else: 
                        processValues(testVal, unitType, valueColumn, valueMultColumn, unitMass)
                    
                else: 
                    if(test == '*'): 
                        thc = currentSample['d9-THC']
                        thcv = currentSample['THCA']  * percentageMulti 
                        total_thc = thc + thcv 
                    
                        print('Tests: %s | Value: %s ' % (test, total_thc))
                        processValues(total_thc, unitType, valueColumn, valueMultColumn, unitMass)
                        
                    if(test == '**'): 
                        cbd = currentSample['CBD']  
                        cbda = currentSample['CBDA'] * percentageMulti 
                        total_cbd = cbd + cbda 

                        print('Tests: %s | Value: %s ' % (test, total_cbd))
                        processValues(total_cbd, unitType, valueColumn, valueMultColumn, unitMass)
                

        if(unitType == 3): 
            currentColumn= 2 + index; 
            #TODO: if there isn't info the nwe can just have one, based on unit
            for i, test in enumerate(dexlueTestsFormatting): 
                
                valueColumn = ws.cell(row=pageLocation+i, column=currentColumn)
                valueColumn.alignment = Alignment(horizontal='left', indent=1)
                
                newBorder = Border(left=thinBorder, right=thinBorder)
                valueColumn.border = combineBorder(valueColumn, newBorder)  
                
                
                if("*" in test): 
                    valueColumn.font = Font(name="Times New Roman", size=9, bold=True) 

                if(test in currentSample.keys()): 
                    testVal = currentSample[test]
                    print('Tests: %s | Value: %s ' % (test, testVal) )

                    if(testVal == 0.0): 
                        valueColumn.value = 'ND' 
                    else: 
                        processValues(testVal, unitType, valueColumn)
                    
                else: 
                    if(test == '*'): 
                        thc = currentSample['d9-THC']
                        thcv = currentSample['THCA'] * percentageMulti 
                        total_thc = thc + thcv 
                    
                        print('Tests: %s | Value: %s ' % (test, total_thc))
                        processValues(total_thc, unitType, valueColumn)
                    
                    if(test == '**'): 
                        cbd = currentSample['CBD']
                        cbda = currentSample['CBDA'] * percentageMulti 
                        total_cbd = cbd + cbda 

                        print('Tests: %s | Value: %s ' % (test, total_cbd))
                        processValues(total_cbd, unitType, valueColumn)
    
        
    pageLocation += len(deluxeHeaders);
        
    return pageLocation  

def insertThcSamples(ws, pageLocation, col, index, units): 
    sampleHeaderName = ws.cell(row=pageLocation, column=col)
    sampleHeaderName.alignment = Alignment(indent=1)
    sampleHeaderName.border = Border(left=thinBorder, right=thinBorder) 

    sampleUnitVal = ws.cell(row=pageLocation+1, column=col)
    sampleUnitVal.alignment = Alignment(indent=1)
    sampleUnitVal.border = Border(left=thinBorder, right=thinBorder) 
    
    sampleHeaderName.value = 'Sample ' + str(index)
    sampleUnitVal.value = units
    




            

def insertBasicTableHeader(ws, pageLocation, sectionJobs, currentSection, start, unitType): 
    
    headerLocations = [
        [1,'Cannabinoids'],
        [6,'Blank'],
        [7,'Recovery'],
        [8,'S₀']
    ]

    unitTypeOptions = [
        ['(mg/g)', '(%)'], 
        ['(mg/mL)', '(%)'], 
        ['(mg/g)', '(mg/unit)'], 
        ['(%)', '(%)']
    ]
    
    for i in range(1,9): 
        bottomBorder = ws.cell(row=pageLocation+1, column=i)
        bottomBorder.border = Border(bottom=doubleBorder)
   
    for currentItem in headerLocations: 
        headerNames = ws.cell(row=pageLocation, column=currentItem[0])
        headerNames.value = currentItem[1]
        headerNames.font = Font(name="Times New Roman", size=9, bold=True)
        headerNames.alignment = Alignment(indent=1) 
        
        if(currentItem[0] >= 6): 
            temp = ws.cell(row=pageLocation+1, column=currentItem[0])
            temp.alignment = Alignment(indent=1) 
            temp.value = ('(%)')
            
            if(currentItem[0] == 6): 
                headerNames.border = Border(left=thinBorder)
                temp.border = Border(left=thinBorder,bottom=doubleBorder)
            else: 
                temp.border = Border(bottom=doubleBorder)
            
        else: 
            headerNames.border = Border(right=thinBorder)
            temp = ws.cell(row=pageLocation+1, column=currentItem[0])
            temp.border = Border(right=thinBorder, bottom=doubleBorder)
        
    counter = 0;
    for index, sample in enumerate(sectionJobs[currentSection]): 
        currentSample = start + index; 
        counter +=1; 
        
        if(unitType in [0,1,2]): 
            if(index == 0): 
                currentColumn = 2; 
            else: 
                currentColumn = 4; 
            #clone basically
            for col in range(2): 
                sampleSection = ws.cell(row=pageLocation, column=currentColumn + col)
                sampleSection.value = f'Sample {currentSample}'
                sampleSection.font = Font(name="Times New Roman", size=9, bold=True)
                sampleSection.alignment = Alignment(indent=1) 
                sampleSection.border = Border(left=thinBorder, right=thinBorder)

                unitSection = ws.cell(row=pageLocation+1, column=currentColumn + col)
                unitSection.value = unitTypeOptions[unitType][col]
                unitSection.alignment = Alignment(indent=1)  
                unitSection.border = Border(left=thinBorder, right=thinBorder, bottom=doubleBorder)
        else: 
            currentColumn = 2 + index; 
            sampleSection = ws.cell(row=pageLocation, column=currentColumn)
            sampleSection.value = f'Sample {currentSample}'
            sampleSection.font = Font(name="Times New Roman", size=9, bold=True)
            sampleSection.alignment = Alignment(indent=1) 
            sampleSection.border = Border(left=thinBorder, right=thinBorder)

            unitSection = ws.cell(row=pageLocation+1, column=currentColumn)
            unitSection.value = unitTypeOptions[unitType][0]
            unitSection.alignment = Alignment(indent=1)  
            unitSection.border = Border(left=thinBorder, right=thinBorder, bottom=doubleBorder) 
            
    start += counter; 

    return pageLocation + 2, start; 

def insertDeluxeTableHeader(ws, pageLocation, sectionJobs, currentSection, start, unitType): 
    
    headerLocations = [
        [1,'Cannabinoids'],
        [6,'Blank'],
        [7,'Recovery'],
        [8,'S₀']
    ]

    unitTypeOptions = [
        ['(mg/g)', '(%)'], 
        ['(mg/mL)', '(%)'], 
        ['(mg/g)', '(mg/unit)'], 
        ['(%)', '(%)']
    ] 





#FIXME: can make this better integrated with the orginal function 
def processValues(threshholdValue, unitType, placementOne, placementTwo = None, mulitpler=None): 
    
    if(unitType in [0,1]): 
        result = formatLeadingValues(threshholdValue); 
        placementOne.value = result 
        
        percentageVal = threshholdValue/10 
        percentageVal = formatLeadingValues(percentageVal) 
        
        placementTwo.value = percentageVal; 

    if(unitType == 2): 
        result = formatLeadingValues(threshholdValue); 
        placementOne.value = result 
        
        unitMassValue = threshholdValue * mulitpler
        unitMassValue = formatLeadingValues(unitMassValue)
        
        placementTwo.value = unitMassValue 
        
        
    if(unitType == 3): 
        result = threshholdValue/10 
        result = formatLeadingValues(result) 
        placementOne.value = result; 
    
    
            
            
def formatLeadingValues(threshholdValue): 
    if(threshholdValue >= 100 ): 
        result = round(threshholdValue, 0)

    if(threshholdValue > 1 and threshholdValue < 100): 
        result = round(threshholdValue, 2)
    
    if(threshholdValue <= 1): 
        result = round(threshholdValue, 3)

    return result; 
            
           

def basicTableFormating(ws, pageLocation): 
     
    for i in range(1, 9): 
        thc = ws.cell(row=pageLocation, column = i); 
        thc.border = Border(top=thinBorder) 

        thca = ws.cell(row=pageLocation+1, column = i)

        total = ws.cell(row=pageLocation+2, column = i)
        total.border = Border(top=thinBorder, bottom=thinBorder)

        d8thc = ws.cell(row=pageLocation+3, column = i); 
        d8thc.border = Border(top=thinBorder, bottom=thinBorder) 

        cbd = ws.cell(row=pageLocation+4, column = i)
        cbda = ws.cell(row=pageLocation+5, column = i) 

        totalcbd = ws.cell(row=pageLocation+6, column=i)
        totalcbd.border = Border(top=thinBorder, bottom=thinBorder)

        cbn = ws.cell(row=pageLocation+7, column=i)
        cbna = ws.cell(row=pageLocation+8, column=i)

        temp5 = ws.cell(row=pageLocation+9, column=i)
        temp5.border = Border(top=thinBorder, bottom=thinBorder); 

        if(i > 1 and i < 7):
            thc.border = Border(top=thinBorder, left=thinBorder)
            thca.border = Border(left=thinBorder)
            total.border = Border(top=thinBorder, bottom=thinBorder, left=thinBorder)
            d8thc.border = Border(top=thinBorder, bottom=thinBorder, left=thinBorder) 
            cbd.border = Border(left=thinBorder)
            cbda.border = Border(left=thinBorder) 
            totalcbd.border = Border(top=thinBorder, bottom=thinBorder, left=thinBorder)
            cbn.border = Border(left=thinBorder)
            cbna.border = Border(left=thinBorder) 
            
        if(i in [2,4,6]): 
            temp5.border = Border(top=thinBorder, bottom=thinBorder, left=thinBorder)  
            
            


def insertThcComment(ws, pageLocation): 
    comments = [
            'Methods: solvent extraction; measured by UPLC-UV, tandem MS, P.I. 1.14 & based on USP monograph 29', 
            'So = standard deviation at zero analyte concentration. MDL generally considered to be 3x So value.', 
            '', 
            'ND = none detected. N/A = not applicable. THC = tetrahydrocannabinol.', 
            '*Total THC = ∆9-THC + (THCA x 0.877 ). **Total CBD = CBD + (CBDA x 0.877).', 
            '', 
            'Material will be held for up to 3 weeks unless alternative arrangements have been made. Sample holding time may vary and is',  
            'dependent on MBL license restrictions.'
        ] 

    for i, comment in enumerate(comments): 
        #print(i, comment)
        temp = ws.cell(row = pageLocation+i, column=1)
        temp.value = comment 
        

    pageLocation+= len(comments) + 3; 
    
    insertSignature(ws, pageLocation, [3,6])
    return pageLocation; 
    

def generateSampleSectionNames(samples, sampleNames, unitType):
    
    if(unitType in [0,1,2]):
        tableTotalSamples = 2
    else: 
        tableTotalSamples = 4; 
     
    currentWord = ''
    headerNames = []
    
    for i, sample in enumerate(samples, 1):
        
        if sample in sampleNames: 
            currentWord += str(i) + ") " +  sampleNames[sample].strip() + " "
        else: 
            currentWord += str(i) + ")" + sample + ' '
        
        if(i % tableTotalSamples == 0): 
            headerNames.append(currentWord)
            currentWord = ''
            
    if(currentWord != ''): 
        headerNames.append(currentWord)

    return headerNames; 

    
    
