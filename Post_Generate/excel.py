

import pickle 
import os 
import math; 

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, borders, Border, Side
from openpyxl.cell.rich_text import TextBlock, CellRichText 
from openpyxl.worksheet.page import PageMargins
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter


from Modules.utilities import * 


#Fonts
defaultFont = Font(name="Times New Roman", size=9)

#Borders 
thinBorder = Side(border_style="thin", color="000000")
doubleBorder = Side(border_style='double', color="000000")

# Create custom number formats
three_decimal_format = '0.000' 
two_decimal_format = "0.00"   # Two decimal places for 1.00-9.99
one_decimal_format = "0.0"    # One decimal place for 10.0-99.9
no_decimal_format = "0"       # No decimal points for anything greater

def significantFiguresConvert(value): 
    if(value >= 100): 
        return int(f'{value:.0f}')
    if(value >=10): 
        return float(f'{value:.1f}')
    if(value >= 1): 
        return float(f'{value:.2f}')
    if(value < 1): 
        return float(f'{value:.2f}') 
    
    return value; 
    
def get_format_for_value(value):
    if value < 1.0:
        return three_decimal_format
    if 1.00 <= value <= 9.99:
        return two_decimal_format
    if 10.0 <= value <= 99.9:
        return one_decimal_format
    else:
        return no_decimal_format

def insertSampleName(ws, row, sampleSection, endCol): 
    print('sample selection: ', sampleSection)
    
    temp = ws.cell(row=row, column=1)
    temp.value = 'Samples: ' + sampleSection
    temp.border = Border(bottom=thinBorder)
   
    ws.merge_cells(start_row=row, start_column=1, end_row=row+1, end_column=endCol)
    temp.alignment = Alignment(wrap_text=True) 
    
    return row + 3; 

def pageSetup(ws):  
    # set the default view to page layout
    ws.sheet_view.view = "pageLayout" 

    # Set the page width to auto
    page_setup = ws.page_setup
    
    #setup page size 
    page_setup.fitToPage = True
    page_setup.fitToHeight = False 
    page_setup.fitToWidth = True
    
    #page margins     
    page_margins = PageMargins()
    page_margins.left = 0.7
    page_margins.right = 0.7
    page_margins.top = 0.75
    page_margins.bottom = 0.75
    page_margins.header = 0.3
    page_margins.footer = 0.3
    
    ws.page_margins = page_margins
    

def formatRows(ws, pageSize, totalPages, maxCol): 
    
    print('Total Pages: ', totalPages)

    totalRows = (pageSize * totalPages) - (8 * (totalPages-1)) 
    print('Total Rows: ', totalRows)

    for row in ws.iter_rows(min_row=1, max_row=totalRows, min_col=1, max_col=maxCol): 
        #print(row)
        for cell in row:
            cell.font = defaultFont 
            

def createFooters(ws, title, jobNumber): 
    
    #Setting up the headers and footers 
    ws.oddHeader.fontName = 'Times New Roman'
    ws.oddHeader.fontSize = 14

    ws.evenHeader.left.font_name = 'Times New Roman'
    ws.evenHeader.left.font_size = 14
    
    ws.oddHeader.left.text  = title + ': &D'
    ws.evenHeader.left.text = title + ': &D' 
    
    ws.oddHeader.right.text  = f"Page &P of &N \n W{jobNumber}"
    ws.evenFooter.right.text = f"Page &P of &N \n W{jobNumber}"
    
    ws.oddFooter.left.text  = '&BT:&B 250 656 1334 \n&BE:&B info@mblabs.com'
    ws.evenFooter.left.text = '&BT:&B 250 656 1334 \n&BE:&B info@mblabs.com' 
    
    ws.oddFooter.center.text = "&B MB Laboratories Ltd.&B \nwww.mblabs.com "
    ws.evenFooter.center.text= "&B MB Laboratories Ltd.&B \nwww.mblabs.com "
    
    ws.oddFooter.right.text = '&BMail:&B PO BOX 2103 Stn Main \n Sidney, B.C, V8L 356'
    ws.evenFooter.right.text ='&BMail:&B PO BOX 2103 Stn Main \n Sidney, B.C, V8L 356'
    
    

def insertClientInfo(ws, clientInfo, column2): 
    
    ws['A1'] = clientInfo['clientName']
        
    if(clientInfo['attn'] is None): 
        ws['A2'] = '*' 
    else: 
        ws['A2'] = clientInfo['attn'] 
        
    ws['A3'] = clientInfo['addy1']
    ws['A4'] = clientInfo['addy2'] + ", " + clientInfo['addy3']
    
    ws['A6'] = 'TEL: ' + clientInfo['tel']
    ws['A7'] =  clientInfo['email']
    
    #ws[column2 + '1'] = "Date: " + clientInfo['date'] + "  (" + clientInfo['time'] + ")" 
    dateCell = ws[column2 + '1'] 
    test_text= "Date: " + clientInfo['date'] + "  (" + clientInfo['time'] + ")" 
    dateCell.value = test_text
    start = 1
    end =  5
    for idx, char in enumerate(test_text, start=1):
        if start <= idx <= end:
            dateCell.font = Font(bold=True)
        else:
            dateCell.font = Font(bold=False)
    
    ws[column2 + '2'] = "Source: " + clientInfo['sampleType1']
    ws[column2 + '3'] = "Type: " + clientInfo['sampleType2']
    ws[column2 + '4'] = "No. of Samples: " + clientInfo['totalSamples']
    ws[column2 + '6'] = "Arrival temp: " + clientInfo['recvTemp']
    ws[column2 + '7'] = "PD: " + clientInfo['payment']
    
    return ws 


def insertNextSectionComment(ws, pageLocation): 
    comment = ws.cell(row=pageLocation, column=1)
    comment.value = 'Continued on next page ....'
    comment.font = Font(bold=True, size=9, name="Times New Roman")


def insertSignature(ws, pageLocation, startColumn): 
    names = [
        'R. Biloduea', 
        'H. Hartmann'
    ]
    postions = [
        'Analytical Chemist:',
        'Sr Analytical Chemist:'
    ]
    
    for i, col in enumerate(startColumn): 
        scientistName = ws.cell(row=pageLocation, column=col) 
        scientistPostion = ws.cell(row=pageLocation+1, column=col)
        
        scientistName.value = names[i]
        scientistPostion.value = postions[i]
        
        for j in range(2): 
            signatureLine = ws.cell(row=pageLocation, column=col+j)
            signatureLine.border = Border(top=thinBorder)