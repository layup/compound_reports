import pickle 
import os 
import math; 

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, borders, Border, Side, NamedStyle
from openpyxl.cell.rich_text import TextBlock, CellRichText 
from openpyxl.worksheet.page import PageMargins
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

from Modules.utilities import * 
from Post_Generate.excel import * 



#TODO: process the if client info is blank 
#TODO: make it an output folder when it goes out 
#TODO: create a bench sheet with all the information 

COMPONENTS = {
    1: 'Abamectin',
    2: 'Acephate',
    3: 'Acequinocyl',
    4: 'Acetamiprid',
    5: 'Aldicarb',
    6: 'Allethrin',
    7: 'Azadirachtin',
    8: 'Azoxystrobin',
    9: 'Benzovindiflupyr',
    10: 'Bifenazate',
    11: 'Bifenthrin',
    12: 'Boscalid',
    13: 'Buprofezin',
    14: 'Carbaryl',
    15: 'Carbofuran',
    16: 'Chlorantraniliprole',
    17: 'Chlorphenapyr',
    18: 'Chlorpyrifos',
    19: 'Clofentezine',
    20: 'Clothianidin',
    21: 'Coumaphos',
    22: 'Cyantraniliprole',
    23: 'Cyfluthrin',
    24: 'Cypermethrin',
    25: 'Cyprodinil',
    26: 'Daminozide',
    27: 'Deltamethrin',
    28: 'Diazinon',
    29: 'Dichlorvos',
    30: 'Dimethoate',
    31: 'Dimethomorph',
    32: 'Dinotefuran',
    33: 'Dodemorph',
    34: 'Endosulfan-alpha',
    35: 'Endosulfan-beta',
    36: 'Endosulfan-sulfate',
    37: 'Ethoprophos',
    38: 'Etofenprox',
    39: 'Etoxazole',
    40: 'Etridiazole',
    41: 'Fenoxycarb',
    42: 'Fenpyroximate',
    43: 'Fensulfothion',
    44: 'Fenthion',
    45: 'Fenvalerate',
    46: 'Fipronil',
    47: 'Flonicamid',
    48: 'Fludioxonil',
    49: 'Fluopyram',
    50: 'Hexythiazox',
    51: 'Imazalil',
    52: 'Imidacloprid',
    53: 'Iprodione',
    54: 'Kinoprene',
    55: 'Kresoxim-methyl',
    56: 'Malathion',
    57: 'Metalaxyl',
    58: 'Methiocarb',
    59: 'Methomyl',
    60: 'Methoprene',
    61: 'Methyl parathion',
    62: 'Mevinphos',
    63: 'MGK-264',
    64: 'Myclobutanil',
    65: 'Naled (Dibrom)',
    66: 'Novaluron',
    67: 'Oxamyl',
    68: 'Paclobutrazol',
    69: 'Permethrin',
    70: 'Phenothrin',
    71: 'Phosmet',
    72: 'Piperonyl butoxide',
    73: 'Pirimicarb',
    74: 'Prallethrin',
    75: 'Propiconazole',
    76: 'Propoxur',
    77: 'Pyraclostrobin',
    78: 'Pyrethrin I',
    79: 'Pyrethrin II',
    80: 'Pyridaben',
    81: 'Quintozene',
    82: 'Resmethrin',
    83: 'Spinetoram',
    84: 'Spinosad',
    85: 'Spirodiclofen',
    86: 'Spiromesifen',
    87: 'Spirotetramat',
    88: 'Spiroxamine',
    89: 'Tebuconazole',
    90: 'Tebufenozide',
    91: 'Teflubenzuron',
    92: 'Tetrachlorvinphos',
    93: 'Tetramethrin',
    94: 'Thiacloprid',
    95: 'Thiamethoxam',
    96: 'Thiophanate-methyl',
    97: 'Trifloxystrobin',
    98: 'Aflatoxin B1',
    99: 'Aflatoxin B2',
    100: 'Aflatoxin G1',
    101: 'Aflatoxin G2',
    102: 'Ochratoxin',
    103: 'Zearalenone',
}


LOQ_VALUES = loadLOQ()


#determine if single/multi and toxin type
#TODO: mkdir for the folders before 
def reportPaths(jobNumbers, sampleTypes, excelFileName): 
    
    results = loadLocations(); 
    path = results['output'];  

    pestPaths = {}
    toxicPaths = {}
    
    textContent = f'Source File: {excelFileName}' 

    for jobNum, value in sampleTypes.items(): 
        
        textFileName = f'{jobNum}_source.txt'
        folderPath = os.path.join(path, jobNum)
        
        if not os.path.exists(folderPath):
            os.makedirs(folderPath)
            textPath = os.path.join(folderPath, textFileName)
            try: 
                with open(textPath, "w") as file:
                    file.write(textContent)
            except: 
                pass; 
            
        pestMultiFileName = 'W' + str(jobNum) + '_pesticides.xlsx'   
        toxicMultiFileName = 'W' + str(jobNum) + '_toxic.xlsx'
        
        pestMultiFilePath =  os.path.join(folderPath, pestMultiFileName)
        toxicMultiFilePath = os.path.join(folderPath, toxicMultiFileName) 
            
        pestSinglePaths = {}
        toxicSinglePaths = {}
        
        for pathType, value2 in value.items(): 
            if(pathType == 'single'):
                for item in value2: 
                    pestSingleFileName = 'W' + str(item) + '_pesticides.xlsx'   
                    toxicSingleFileName = 'W' + str(item) + '_toxic.xlsx' 
                    
                    pestSingleFilePath =  os.path.join(folderPath, pestSingleFileName) 
                    toxicSingleFilePath = os.path.join(folderPath, toxicSingleFileName) 
                    
                    pestSinglePaths[item] = pestSingleFilePath
                    toxicSinglePaths[item] = toxicSingleFilePath
                
        pestPaths[jobNum] = {
            'multi': pestMultiFilePath, 
            'single': pestSinglePaths
        }
        toxicPaths[jobNum] = {
            'multi':toxicMultiFilePath,
            'single': toxicSinglePaths 
        }
        
    return pestPaths, toxicPaths  

    
def generatePestReport(jobNumbers, clientInfo, sampleNames, sampleInfo, sampleData, recoveryValues, excelFileName):
    print('****Generating Pesticdes Options')
    print('***Job Numbers: ', jobNumbers)
    print('***Recovery Values') 
    for key, value in recoveryValues.items(): 
        print(key, value)
    #print('Sample Data: ', sampleData) 
   
    sampleTypes = {}
    reportType = {}
    unitType = {}

    pesticidesReport = 0 
    toxinsReport = 1
    bothReports = 2 

    for job in jobNumbers: 
        single = []
        multi = []
        for key, value in sampleInfo.items(): 
   
            if job in key:
                if value[3] == 'Multi': 
                    multi.append(key)
                else: 
                    single.append(key)
                    
                if 'Pesticides' in value[2]: 
                    reportType[job] = pesticidesReport 
                
                if 'Toxins' in value[2]: 
                    reportType[job] = toxinsReport
                
                if 'Both' in value[2]: 
                    reportType[job] = bothReports
                    
                unitType[job] = value[1] 
        
        sampleTypes[job] = {
            'single':single, 
            'multi': multi
        }

    print('**Single/Multi Reports')  
    for jobNumber, sampleType in sampleTypes.items():         
        print(jobNumber, sampleType)
    
    pestPaths, toxicPaths = reportPaths(jobNumbers, sampleTypes, excelFileName)

    #print('**Pesticdes Paths')
    #for key,value in pestPaths.items(): 
    #    print(key, value)

    #print('**Toxins Paths')
    #for key, value in toxicPaths.items(): 
    #    print(key, value)

    
    for jobNum, report in reportType.items(): 
        print('**Current Job Being Created: ', jobNum)
        print('**Report Type: ', report); 
        
        #TODO: both     
        if(report == bothReports): 
            for key, value in sampleTypes[jobNum].items(): 
                print('KEY: ', key, ' value: ', value); 
                if(key == 'single'): 
                    #createPestReport(jobNum, pestPaths[jobNum]['multi'], clientInfo[jobNum],)
                    print('Single Pest Reports: ', value)
                    for singleReport in value: 
                        singlePath = pestPaths[jobNum]['single'][singleReport]
                        temp = [singleReport]
                        print('TEMP: ', temp); 
                        createPestReport(jobNum, singlePath, clientInfo[jobNum], temp, sampleNames[jobNum], sampleData, recoveryValues, unitType[jobNum])
                        createToxinReport(jobNum, singlePath, clientInfo[jobNum], singleReport, sampleNames[jobNum], sampleData, recoveryValues, unitType[jobNum])
                    
                else: 
                    if(len(value) != 0): 
                        print('Multiple Pest Reports: ', value) 
                        createPestReport(jobNum, pestPaths[jobNum]['multi'], clientInfo[jobNum], value, sampleNames[jobNum], sampleData, recoveryValues, unitType[jobNum])
                        createToxinReport(jobNum, toxicPaths[jobNum]['multi'], clientInfo[jobNum], value, sampleNames[jobNum], sampleData, recoveryValues, unitType[jobNum])
            
        if(report == pesticidesReport): 
            #createPestReport(jobNum, clientInfo[jobNum], sampleData, recoveryValues) 
            
            for key, value in sampleTypes[jobNum].items(): 
                print('KEY: ', key, ' value: ', value); 
                if(key == 'single'): 
                    #createPestReport(jobNum, pestPaths[jobNum]['multi'], clientInfo[jobNum],)
                    print('Single Pest Reports: ', value)
                    for singleReport in value: 
                        singlePath = pestPaths[jobNum]['single'][singleReport]
                        temp = [singleReport]
                        print('TEMP: ', temp); 
                        createPestReport(jobNum, singlePath, clientInfo[jobNum], temp, sampleNames[jobNum], sampleData, recoveryValues, unitType[jobNum])
                    
                else: 
                    if(len(value) != 0): 
                        print('Multiple Pest Reports: ', value) 
                        createPestReport(jobNum, pestPaths[jobNum]['multi'], clientInfo[jobNum], value, sampleNames[jobNum], sampleData, recoveryValues, unitType[jobNum])
                    
            
        if(report == toxinsReport): 
            for key, value in sampleTypes[jobNum].items(): 
                print(key,value)
                if(key == 'single'): 
                    for singleReport in value: 
                        singlePath = toxicPaths[jobNum]['single'][singleReport]
                        createToxinReport(jobNum, singlePath, clientInfo[jobNum], singleReport, sampleNames[jobNum], sampleData, recoveryValues, unitType[jobNum])
                        #createPestReport(jobNum, singlePath, clientInfo[jobNum], singleReport, sampleNames[jobNum], sampleData, recoveryValues, unitType[jobNum])
                        
                else: 
                    print('Multiple Toxic Reports: ', value) 
                    createToxinReport(jobNum, toxicPaths[jobNum]['multi'], clientInfo[jobNum], value, sampleNames[jobNum], sampleData, recoveryValues, unitType[jobNum])
                



def createPestReport(jobNum, outputPath, clientInfo, samples, sampleNames, sampleData, recoveryValues, unitType ): 
    print('****Creating Pest Reports: ', jobNum)
    print('**Samples: ', samples)
    print('**Sample Names: ', sampleNames)
    print('*Output Path: ', outputPath)
    print('*unitType', unitType)
    #print(sampleData)
    #print('*Client Info: ', clientInfo)
    
    wb = Workbook()
    ws = wb.active

    pageSetup(ws);    
    headerTitle = 'Pesticides Report'
    createFooters(ws, headerTitle, jobNum);  
    ws = insertClientInfo(ws, clientInfo, 'E')
 
    for i in range(1, 10): 
        columnLetter = get_column_letter(i)
        ws.column_dimensions[columnLetter].width = 11
        
        if(i == 1): 
            ws.column_dimensions[columnLetter].width = 5
            
        if(i == 2): 
            ws.column_dimensions[columnLetter].width = 16
        
    ws.print_title_rows = '1:8' # the first two rows
    
    pageSize = 56; 
    pageLocation = 9; 
    currentPage = 1; 
   
    totalSamples = len(samples)
    totalSections = math.ceil(totalSamples/4)
    totalPages = totalSections * 3 
    maxCols = 9 
    
    print('Total Samples: ', totalSamples, 'Total Sections: ', totalSections)

    formatRows(ws, pageSize, totalPages, maxCols)

    sectionJobs = []; 
    currentSection = 0; 
       
    if(totalSamples <= 4): 
        sectionJobs.append(samples) 
    else: 
        tempSection = []
        for i, sample in enumerate(samples, start=1): 
            tempSection.append(sample)
            if(i % 4 == 0): 
                sectionJobs.append(tempSection)
                tempSection = []
                
        if(len(tempSection) != 0): 
            sectionJobs.append(tempSection);
    
    print('Sections: ', sectionJobs)
    
    totalSections = len(sectionJobs)
    headerNames = generateSampleSectionNames(samples, sampleNames)
        
    print("**Headernames")
    for headerName in headerNames: 
        print(headerName)

    headerCounter = 1; 
    for i, section in enumerate(sectionJobs, 1): 

        pageLocation = insertSampleName(ws, pageLocation, headerNames[currentSection],9)
        pageLocation = insertTableHeader(ws, pageLocation, sectionJobs, currentSection, headerCounter, unitType)
        pageLocation = insertPestTable(ws, pageLocation, 0, 38, recoveryValues, sampleData, sectionJobs, currentSection, unitType)
        pageLocation += 2;  
        insertNextSectionComment(ws, pageLocation)
        
        currentPage += 1; 
        pageLocation = ((currentPage-1) * pageSize) - (8 * (currentPage-2)) + 1; 
        
        pageLocation = insertTableHeader(ws, pageLocation, sectionJobs, currentSection, headerCounter, unitType)
        pageLocation = insertPestTable(ws, pageLocation, 38, 79, recoveryValues, sampleData, sectionJobs, currentSection, unitType) 
        pageLocation += 2;  
        insertNextSectionComment(ws, pageLocation)

        currentPage += 1; 
        pageLocation = ((currentPage-1) * pageSize) - (8 * (currentPage-2)) + 1; 
        
        pageLocation = insertTableHeader(ws, pageLocation, sectionJobs, currentSection, headerCounter, unitType)
        pageLocation = insertPestTable(ws, pageLocation, 79, 97, recoveryValues, sampleData, sectionJobs, currentSection, unitType)
        pageLocation += 1; 
        
        if(i == totalSections): 
            insertPestComment(ws, pageLocation)
        else: 
            insertNextSectionComment(ws, pageLocation)
            currentSection +=1; 
            currentPage+=1; 
            pageLocation = ((currentPage-1) * pageSize) - (8 * (currentPage-2)) + 1; 
            headerCounter += 4; 
            
    
    wb.save(outputPath)


def createToxinReport(jobNum, outputPath, clientInfo, samples, sampleNames, sampleData, recoveryValues, unitType ): 
    print('****Creating Toxic Reports: ', jobNum)
    print('**Samples: ', samples)
    print('**Sample Names: ', sampleNames)
    print('*Output Path: ', outputPath)
    print('*unitType', unitType)
    #print(sampleData)
    #print('*Client Info: ', clientInfo)
    
    wb = Workbook()
    ws = wb.active

    pageSetup(ws);    
    headerTitle = 'Toxins Report'
    createFooters(ws, headerTitle, jobNum); 
    
    #insert client names 
    ws = insertClientInfo(ws, clientInfo, 'E')
 
    for i in range(1, 10): 
        columnLetter = get_column_letter(i)
        
        ws.column_dimensions[columnLetter].width = 11
        
        if(i == 1): 
            ws.column_dimensions[columnLetter].width = 5
            
        if(i == 2): 
            ws.column_dimensions[columnLetter].width = 16
        
    #the first two rows at top of each page 
    ws.print_title_rows = '1:8' 
    
    pageSize = 56; 
    pageLocation = 9; 
    currentPage = 1; 
    
    totalSamples = len(samples)
    totalSections = math.ceil(totalSamples/4)
    totalPages = totalSections * 3 
    maxCols = 9; 
    
    print('Total Samples: ', totalSamples, 'Total Sections: ', totalSections)

    formatRows(ws, pageSize, totalPages, maxCols)

    sectionJobs = []; 
    currentSection = 0; 
       
    if(totalSamples <= 4): 
        sectionJobs.append(samples) 
    else: 
        tempSection = []
        for i, sample in enumerate(samples, start=1): 
            
            tempSection.append(sample)
            if(i % 4 == 0): 
                sectionJobs.append(tempSection)
                tempSection = []
                
        if(len(tempSection) != 0): 
            sectionJobs.append(tempSection);
    
    print('Sections: ', sectionJobs)
    
    totalSections = len(sectionJobs)
    headerNames = generateSampleSectionNames(samples, sampleNames)
        
    print("**Headernames")
    for headerName in headerNames: 
        print(headerName)

    headerCounter = 1; 
    
    for i, section in enumerate(sectionJobs, 1):  
        pageLocation = insertSampleName(ws, pageLocation, headerNames[currentSection], 9)
        pageLocation = insertTableHeader(ws, pageLocation, sectionJobs, currentSection, headerCounter, unitType) 
        pageLocation = insertToxicTable(ws, pageLocation, recoveryValues, sampleData, sectionJobs, currentSection, unitType)
        pageLocation +=1; 

        if(i == totalSections): 
            pageLocation +=1; 
            insertToxicCommets(ws,pageLocation)
            break;  
        
        currentSection += 1;  
        headerCounter += 4;  
        if(i+1 == totalSections): 
            if(i % 3 != 1): 
                pageLocation+=1; 
                insertNextSectionComment(ws, pageLocation)    
                currentPage+=1; 
                pageLocation = ((currentPage-1) * pageSize) - (8 * (currentPage-2)) + 1;  
        else: 
            if(i % 3 == 0): 
                pageLocation+=1; 
                insertNextSectionComment(ws, pageLocation)    
                currentPage+=1; 
                pageLocation = ((currentPage-1) * pageSize) - (8 * (currentPage-2)) + 1; 
                
    wb.save(outputPath)


def insertTableHeader(ws, pageLocation, sectionJobs, currentSection, start, unitType): 

    if unitType == 'Bud': 
        LOQ = 'LOQ (Bud)'    
    if unitType == 'Oil':
        LOQ = 'LOQ (Oil)'
    if unitType == 'Paper': 
        LOQ = 'LOQ (Paper)'
    
    headerLocations = [
        [2, 'Analyte'],
        [7, LOQ, '(ng/g)'],
        [8, 'Blank', '(ng/g)'],
        [9, 'Recovery','(%)']
    ]
    
    for i in range(1,10): 
        bottomBorder = ws.cell(row=pageLocation+1, column=i)
        bottomBorder.border = Border(bottom=doubleBorder) 

    for currentItem in headerLocations: 
        headerNames = ws.cell(row=pageLocation, column=currentItem[0])
        headerNames.value = currentItem[1]
        headerNames.font = Font(name="Times New Roman", size=9, bold=True)
        headerNames.alignment = Alignment(indent=1) 
        
        if(len(currentItem) == 3): 
            temp = ws.cell(row=pageLocation+1, column=currentItem[0])
            temp.alignment = Alignment(indent=1) 
            temp.value = currentItem[2]

            if(currentItem[0] < 9): 
                headerNames.border = Border(left=thinBorder, right=thinBorder)
                temp.border = Border(left=thinBorder, right=thinBorder,bottom=doubleBorder)
        else: 
            headerNames.border = Border(left=thinBorder, right=thinBorder)
            temp = ws.cell(row=pageLocation+1, column=currentItem[0])
            temp.border = Border(left=thinBorder, right=thinBorder, bottom=doubleBorder)
   
    for index, sample in enumerate(sectionJobs[currentSection]): 
        currentColumn = 3 + index; 
        currentSample = start + index; 
        
        sampleSection = ws.cell(row=pageLocation, column=currentColumn)
        sampleSection.value = f'Sample {currentSample}'
        sampleSection.font = Font(name="Times New Roman", size=9, bold=True)
        sampleSection.alignment = Alignment(indent=1) 
        sampleSection.border = Border(left=thinBorder, right=thinBorder)

        unitSection = ws.cell(row=pageLocation+1, column=currentColumn)
        unitSection.value = '(ng/g)'
        unitSection.alignment = Alignment(indent=1)  
        unitSection.border = Border(left=thinBorder, right=thinBorder, bottom=doubleBorder)
        
    return pageLocation + 2; 
    
    #return pageLocation+2; 

def insertPestTable(ws, pageLocation, startingRow, endingRow, recoveryValues, sampleData, sectionJobs,currentSection, unitType): 
    
    tempRow = pageLocation
    
    for index in range(startingRow, endingRow): 
        currentComponet = COMPONENTS[index+1]  
        currentRow = tempRow
        tempRow+=1; 
        
        numberLocation = ws.cell(row=currentRow, column=1); 
        componentLocation = ws.cell(row=currentRow, column=2)
        loqLocation = ws.cell(row=currentRow, column=7)
        blankLocation = ws.cell(row=currentRow, column=8)
        recoveryLocation = ws.cell(row=currentRow, column=9)
        
        numberLocation.value = index+1
        
        componentLocation.value = currentComponet
        componentLocation.alignment = Alignment(horizontal='left', indent=1)
       
        loqValue = insertLOQ(currentComponet, unitType) 
        loqLocation.value = loqValue
        loqLocation.alignment =  Alignment(horizontal='left', indent=1)

        try: 
            number_format = get_format_for_value(loqValue)
            loqLocation.number_format = number_format 
        except: 
            print('could not convert the LOQ value')

        blankLocation.value = 'ND'
        blankLocation.alignment = Alignment(indent=1)
        
        recoveryValue = recoveryValues[currentComponet]
        recoveryLocation.value = recoveryValue
        recoveryLocation.alignment = Alignment(horizontal='left', indent=1)   

        try: 
            number_format = get_format_for_value(float(recoveryValue))
            recoveryValue.number_format = number_format 
        except: 
            newRecoveryValue = significantFiguresConvert(recoveryValue)
            recoveryLocation.value = newRecoveryValue 
        
        for column, cell in enumerate(ws[currentRow][0:9]): 
            if(column in [0,1,5,6,7]): 
                cell.border = Border(bottom=thinBorder, right=thinBorder)
            else: 
                cell.border = Border(bottom=thinBorder)
                
        #INSERTING VALUES 
        for j, sample in enumerate(sectionJobs[currentSection]): 
            sampleCell = ws.cell(row=currentRow, column=3+j)
            sampleValue = sampleData[sample][index][1]
            sampleCell.value = sampleValue
            
            if(sampleValue != 'ND'): 
                try: 
                    number_format = get_format_for_value(sampleValue)
                    sampleCell.number_format = number_format 
                except: 
                    pass; 
            
            sampleCell.alignment = Alignment(horizontal='left', indent=1)     
            sampleCell.border = Border(left=thinBorder, right=thinBorder, bottom=thinBorder) 
            
    return pageLocation + (endingRow-startingRow); 

def insertToxicTable(ws, pageLocation, recoveryValues, sampleData, sectionJobs, currentSection, unitType): 
    tempRow = pageLocation; 

    startingIndex = 97; 
    
    for index in range(1,7): 
        currentComponet = COMPONENTS[startingIndex+index]; 
        currentRow = tempRow; 
        tempRow+=1; 

        numberLocation = ws.cell(row=currentRow, column=1); 
        componentLocation = ws.cell(row=currentRow, column=2)
        loqLocation = ws.cell(row=currentRow, column=7)
        blankLocation = ws.cell(row=currentRow, column=8)
        recoveryLocation = ws.cell(row=currentRow, column=9)
        
        numberLocation.value = index+1
        
        if(currentComponet == 'Ochratoxin'): 
            componentLocation.value = 'Ochratoxin A'
        else: 
            componentLocation.value = currentComponet
            
        componentLocation.alignment = Alignment(horizontal='left', indent=1)
       
        loqValue = insertLOQ(currentComponet, unitType) 
        loqLocation.value = loqValue
        loqLocation.alignment =  Alignment(horizontal='left', indent=1)

        try: 
            number_format = get_format_for_value(loqValue)
            loqLocation.number_format = number_format 
        except: 
            print('could not convert the LOQ value')

        
        blankLocation.value = 'ND'
        blankLocation.alignment = Alignment(indent=1)
        
        recoveryValue = recoveryValues[currentComponet]
        recoveryLocation.value = recoveryValue
        recoveryLocation.alignment = Alignment(horizontal='left', indent=1)   

        for column, cell in enumerate(ws[currentRow][0:9]): 
            if(column in [0,1,5,6,7]): 
                cell.border = Border(bottom=thinBorder, right=thinBorder)
            else: 
                cell.border = Border(bottom=thinBorder)
                
        #print(f'**index: {index+1} Component: {COMPONENTS[index+1]}')
        #INSERTING VALUES 
        for j, sample in enumerate(sectionJobs[currentSection]): 
            sampleCell = ws.cell(row=currentRow, column=3+j)
            sampleValue = sampleData[sample][startingIndex+index][1]
            sampleCell.value = sampleValue
        
            if(sampleValue != 'ND'): 
                try: 
                    number_format = get_format_for_value(sampleValue)
                    sampleCell.number_format = number_format 
                except: 
                    pass; 
       
        
            sampleCell.alignment = Alignment(horizontal='left', indent=1)     
            sampleCell.border = Border(left=thinBorder, right=thinBorder, bottom=thinBorder)
            
    pageLocation += 6; 
    return pageLocation; 


#0 = BUD 
#1 = OIL
#2 = Paper
def insertLOQ(compoundName, unitType): 
    if unitType == 'Bud': 
        return LOQ_VALUES[compoundName][0]
    if unitType == 'Oil':
        return LOQ_VALUES[compoundName][1]
    if unitType == 'Paper': 
        return LOQ_VALUES[compoundName][2]
    
    return ''

def insertPestComment(ws, pageLocation): 
    comments = [
        '*Analysis includes all 97 target compounds on the Health Canada Mandatory List Aug 2019', 
        '**Trace = presence & identity of compound verified, value below limit of quantification ', 
        '', 
        "As per international standards, all observed values are reported even if they are below LOQ's. LOQ or MDL's are interpretative ", 
        '& given as guidance only & do not affect reported results.', 
        '', 
        'Method: Sample is solvent extracted, then cleaned using SPE (QuEChERS) methods. Multiresidue analysis is carried out using',  
        'Procedure ref AOAC 2007.01; USP <561><565>, EU 2.0813. Methods fully validated. '
    ] 

    for i, comment in enumerate(comments): 
        #print(i, comment)
        temp = ws.cell(row = pageLocation+i, column=1)
        temp.value = comment 
        
    pageLocation+= len(comments) + 2; 
    
    insertSignature(ws, pageLocation, [3,7])
    pageLocation+=3; 
    
    return pageLocation; 

def insertToxicCommets(ws, pageLocation): 
    comments = [
        'Method: Sample is solvent extracted, then cleaned using SPE (QuEChERS) methods. Multiresidue analysis is carried out using',
        'UPLC-ESI-MS/MS/APCI & GC-MS: SPME. Detection of compounds meet or exceed HC requirements.', 
        'Procedure ref AOAC 2007.01; USP <561><565>, EU 2.0813. methods fully validated.',
        '', 
        'LOQ = Limit of quantification', 
        'ND = none detected n/a = not applicable', 
        'ppb = parts per billion (ng/g)', 
        '',
        'Mycotoxin - Maximum Tolerance Levels -CFIA FAO Food & Nutrition Paper 64, 1997',
        '                         CFIA - Fact Sheet - Mycotoxins LL Charmley & HL Trenholm May 2010', 
        '', 
        'Afalatoxin:             15 ppb           nut products            Cananda', 
        '                                   20 ppb           all foods                   USA',
        '',
        'Ochratoxin A:        20 ppb           Cannabis                 Health Cananda',
        '                                   5-10 ppb       food & spices        EU', 
        '', 
        'Zearalenone:           20-400 ppb  Cannabis                 Health Cananda'
        ''
    ]

    for i, comment in enumerate(comments): 
        #print(i, comment)
        temp = ws.cell(row = pageLocation+i, column=1)
        temp.value = comment 
        
    pageLocation+= len(comments) + 3; 
    
    insertSignature(ws, pageLocation, [3,7])
    pageLocation+=3; 

def generateSampleSectionNames(samples, sampleNames): 
    currentWord = ''
    headerNames = []
    
    for i, sample in enumerate(samples, 1):
        
        if sample in sampleNames: 
            currentWord += str(i) + ") " +  sampleNames[sample].strip() + " "
        else: 
            currentWord += str(i) + ")" + sample + ' '
        
        if(i % 4 == 0): 
            headerNames.append(currentWord)
            currentWord = ''
            
    if(currentWord != ''): 
        headerNames.append(currentWord)

    return headerNames; 