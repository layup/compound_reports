
import re; 
import pickle; 
import os; 
from copy import copy; 

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from Modules.utilities import * 


def scanForTXTFolders(jobNum): 
    #print('jobnumber: ', jobNum)

    fileLocationsDictonary = loadLocations()
    TXTLocation = fileLocationsDictonary['txtLocation']
    
    locationsObject = os.scandir(TXTLocation)
    
    txtFolderLocations = [] 
    
    for entry in locationsObject: 
        if(entry.is_dir()):
            if(re.match('^TXT-[a-zA-Z]{3}$', entry.name)):

                txtFolderLocations.append(os.path.join(TXTLocation, entry.name))
            
    
    #print(txtFolderLocations)
    locationsObject.close()
    return processTXTFolders(jobNum, txtFolderLocations)
  
 

def processTXTFolders(jobNum, locations):
    
    fileName = "W" + jobNum + ".TXT"
    
    #print("list")
    #print(locations)
   
    for i in range(len(locations)): 
        tempLocationObject = os.scandir(locations[i]) 

        for entry in tempLocationObject: 
            if(entry.is_file()): 
                if(re.match(fileName, entry.name)): 
                    print("TXT File found")
                    #print(entry.name)
                    tempLocationObject.close()
                    return os.path.join(locations[i], entry.name)
        
        tempLocationObject.close()
    #TODO: return a blank user information 
    #can just clone the clientInfoDict somewhere and send it back 
    #print("No Job Number Matches")
    return None; 

        
def processClientInfo(jobNum, fileLocation):
    
    clientInfoDict = {
        'clientName': '', 
        'date': '', 
        'time': '', 
        'attn': '', 
        'addy1': '', 
        'addy2': '', 
        'addy3': '', 
        'sampleType1': '', 
        'sampleType2': '', 
        'totalSamples': '', 
        'recvTemp': '', 
        'tel': '', 
        'email': '', 
        'fax': '', 
        'payment': ''
    }
    
    #grab the file names 
    sampleNames = {}
    

    #have the information about the file, what kind of reports and etc 
    sampleTests = {}

    sampleCounter = 0; 
    prevLine = [0, ""]
    prevLineHelper = [0, ""]
    
    if(fileLocation == None): 
        return clientInfoDict, sampleNames; 
        #return clientInfoDict, sampleNames, sampleTests; 
    
    with open(fileLocation) as file: 
    
        for lineLocation, line in enumerate(file, 0):

            if(prevLine[0]+1 == prevLineHelper[0]):
                prevLine[0] = copy(prevLineHelper[0])
                prevLine[1] = copy(prevLineHelper[1])
                prevLineHelper[0] = copy(int(lineLocation))
                prevLineHelper[1] = copy(line)
            else: 
                prevLineHelper[0] = copy(int(lineLocation))
                prevLineHelper[1] = copy(line)
            
            #print('PrevLine: ', prevLine[0], prevLine[1])
            #print('PrevLineHelper: ', prevLineHelper[0], prevLineHelper[1])
            #print('currentLine: ', lineLocation, line)
                    
            if(lineLocation == 1): 
                clientInfoDict['clientName'] = line[0:54].strip()
                clientInfoDict['date'] = line[50:(54+7)].strip()
                clientInfoDict['time'] = line[66:71].strip()
                
            if(lineLocation == 2): 
                clientInfoDict['sampleType1'] = line[54:].strip()
                
                if "*" in line: 
                    clientInfoDict['attn'] = line[:54].strip()
                else: 
                    clientInfoDict['addy1'] = line[:54].strip()
                
            if(lineLocation == 3): 
                clientInfoDict['sampleType2'] = line[54:].strip()
                
                if(clientInfoDict['attn'] != ''):
                    clientInfoDict['addy1'] = line[:60].strip()
                else: 
                    clientInfoDict['addy2'] = line[:60].strip()
            
            if(lineLocation == 4): 
                clientInfoDict['totalSamples'] = line[60:].strip()
                
                if(clientInfoDict['attn'] != ''):
                    clientInfoDict['addy2'] = line[:60].strip()
                else: 
                    clientInfoDict['addy3'] = line[:60].strip() 
                    
            if(lineLocation == 5): 
                if(clientInfoDict['attn'] and clientInfoDict['addy2']): 
                    clientInfoDict['addy3'] = line[:60].strip()
                else: 
                    clientInfoDict['tel'] = line[26:50].strip()

                    try: 
                        clientInfoDict['recvTemp'] = line[71:].strip()
                    except:
                        print('No recv temp avaliable')
                        
            if(lineLocation == 6): 
                clientInfoDict['tel'] = line[26:50].strip() 
                clientInfoDict['recvTemp'] = line[71:].strip()
            
            if(lineLocation == 7): 
                clientInfoDict['fax'] = line[26:].strip()
                
            if(lineLocation == 8): 
                
                try: 
                    foundEmail = re.search('([A-Za-z0-9]+[.-_])*[A-Za-z0-9]+@[A-Za-z0-9-]+(\.[A-Z|a-z]{2,})+', line).group()
                    if(foundEmail): 
                        clientInfoDict['email'] = foundEmail; 
                except:
                    print("email error")
                
                if("pd" in line.lower()): 
                    clientInfoDict['payment'] = line[51:].strip()
                     
            
            if(lineLocation > 35 and len(line) > 0): 
               
                if(sampleCounter != int(clientInfoDict['totalSamples']) ):

                    try: 
                        sampleMatch = re.search('(?<=\s[0-9]).*', line).group()
                        if(sampleMatch): 
                            sampleName = str(jobNum) + '-' + str(sampleCounter+1)
                            sampleNames[sampleName] = sampleMatch.strip()
                            sampleCounter+=1; 
                           
                        #TODO: add something to get the - afterwords 
                    except: 
                        pass
                #find the report information that does with corrisponding thing                
                if(re.search('(?<=\s-\s).*', line)):
                    prevSampleName = str(jobNum) + "-" + str(sampleCounter-1)
                    #print('CURRENT: ', line)
                    #print('PREV: ', prevLine[1])
                    currentTestsCheck = re.search('(?<=\s-\s).*', line)
                    prevSampleMatchCheck = re.search('(?<=\s[0-9]).*', prevLine[1])
                    prevSampleTestsCheck = re.search('(?<=\s-\s).*', prevLine[1])
                    #sampleTests[prevSampleName] = currentTestsCheck.group()
                    #not could be apart of the string name longer 
                    
                    #add to most recent sample
                    if currentTestsCheck:
                        print('current is a test:', currentTestsCheck.group())
                        if prevSampleTestsCheck:
                            sampleTests[prevSampleName] = sampleTests.get(prevSampleName, '') + ", " + currentTestsCheck.group()
                        else:
                            sampleTests[prevSampleName] = sampleTests.get(prevSampleName, '') + currentTestsCheck.group()
                        
                    #Prev sample name 
                    if(prevSampleMatchCheck):
                        print('prev was a sampleName')
                        #print(sampleName[prevSampleName]) #doesnt work 
                        pass; 
                       
                    #append onto them 

                        
                    #TODO: solve this later, add previous name onto current name sample 
                    if((not bool(prevSampleMatchCheck)) and not( bool(prevSampleTestsCheck))): 
                        print('prev was apart of the name yo')
                        
                        
            
            #print('---------------------------') 
                
                    
    file.close()
    
    #print(sampleTests)
    #process tyhe sampleTests 
    for key,value in sampleTests.items():
        
        testLists = [x.strip() for x in value.split(',')]
        sampleTests[key] = testLists
           
    
    #print(clientInfoDict)
    #print(sampleNames)
    #print(sampleTests)
    return clientInfoDict, sampleNames; 


def scanTHC(fileLocation): 
    #TODO: if there is no 1/10 then use the 1/100 insted 
    
    wb = load_workbook(fileLocation) 
    ws = wb.active

    sampleJobLocations = [] 
    sampleRows = []
    
    jobNumbers = []
    sampleNumbers = []
    dilutedSamples = []
    
    sampleNamesPattern =  r"\d{6}-\d{1,2}$" 
    sampleNamesPattern2 = r"\d{6}-\d(?!\/\d+)"

    #sample names and locations
    #TODO: something about 1/100 dillution 
    for cell in ws['AJ']:         
        currentValue = str(cell.value)
        if re.match(sampleNamesPattern2, currentValue): 
            print('MATCH: ', currentValue)
            sampleJobLocations.append([currentValue, cell.row]) 
            sampleRows.append(cell.row)
            
            currentJob = currentValue[:6]
            
            if(currentJob not in jobNumbers): 
                jobNumbers.append(currentJob)
               
            if('1/100' not in currentValue and currentValue not in sampleNumbers): 
                sampleNumbers.append(currentValue)
            
            if('1/100' in currentValue and currentValue not in dilutedSamples):
                dilutedSamples.append(currentValue)
                
                
    recoveryRows = []
    recoveryValues = {}
    stdconc = 2 

    for cell in ws['AL']: 
        if(cell.value == stdconc): 
            recoveryRows.append(cell.row)

    for row in recoveryRows: 
        recovery = ws['EH' + str(row)].value 
        test = ws['BK' + str(row)].value  
        
        if test not in recoveryValues: 
            recoveryValues[test] = recovery 
    
    
    currentName = ''
    currentTests = []
    currentUnits = []
    temp = {}
    sampleData = {}
    sampleData2 = {}

    for sample in sampleJobLocations: 
 
        if(currentName == ''): 
            currentName = sample[0]
        
        if(currentName == sample[0]): 
            currentRow = str(sample[1])

            test = ws['BK' + currentRow].value 
            units = ws['BX' + currentRow].value 


            if(units == None): 
                units = 0; 
            
            currentTests.append(test)
            currentUnits.append(float(units)/1000)
            temp[test] = float(units)/1000
            
        else: 
            sampleData[currentName] = {
                'tests': currentTests, 
                'values': currentUnits 
            }
            sampleData2[currentName] = temp; 

            currentName = sample[0]
            currentTests = [] 
            currentUnits = []
            temp = {}

            #print(sample[0], test, recovery, float(units)/1000, float(units)/10000)
    
    sampleData[currentName] = {
        'tests': currentTests, 
        'values': currentUnits 
    } 
    sampleData2[currentName] = temp; 
    
    wb.close()
            
    print('**Sample Data')
    for key, value in sampleData2.items(): 
        print(key, value )
        
    newData = {}

    for sampleName in dilutedSamples: 
        newName = sampleName.replace(' 1/100', '')
        
        if(newName not in sampleNumbers): 
            newData[newName] = sampleData2[sampleName]
    
    keys = newData.keys()  # Get a view of the keys
    key_list = list(keys)
    difference = set(sampleNumbers) - set(key_list)
     
    for sampleName in difference: 
        newData[sampleName] = sampleData2[sampleName]
        
    print('**NEW SAMPLE DATA')
    for key, value in newData.items(): 
        print(key, value )
        
    return jobNumbers, recoveryValues, newData; 
        
        

values = [
    "Abamectin", "Acephate", "Acequinocyl", "Acetamiprid", "Aldicarb", "Allethrin", "Azadirachtin", "Azoxystrobin",
    "Benzovindiflupyr", "Bifenazate", "Bifenthrin", "Boscalid", "Buprofezin", "Carbaryl", "Carbofuran",
    "Chlorantraniliprole", "Chlorphenapyr", "Chlorpyrifos", "Clofentezine", "Clothianidin", "Coumaphos",
    "Cyantraniliprole", "Cyfluthrin", "Cypermethrin", "Cyprodinil", "Daminozide", "Deltamethrin", "Diazinon",
    "Dichlorvos", "Dimethoate", "Dimethomorph", "Dinotefuran", "Dodemorph", "Endosulfan-alpha", "Endosulfan-beta",
    "Endosulfan-sulfate", "Ethoprophos", "Etofenprox", "Etoxazole", "Etridiazole", "Fenoxycarb", "Fenpyroximate",
    "Fensulfothion", "Fenthion", "Fenvalerate", "Fipronil", "Flonicamid", "Fludioxonil", "Fluopyram",
    "Hexythiazox", "Imazalil", "Imidacloprid", "Iprodione", "Kinoprene", "Kresoxim-methyl", "Malathion",
    "Metalaxyl", "Methiocarb", "Methomyl", "Methoprene", "Methyl parathion", "Mevinphos", "MGK-264", "Myclobutanil",
    "Naled (Dibrom)", "Novaluron", "Oxamyl", "Paclobutrazol", "Permethrin", "Phenothrin", "Phosmet",
    "Piperonyl butoxide", "Pirimicarb", "Prallethrin", "Propiconazole", "Propoxur", "Pyraclostrobin",
    "Pyrethrin I", "Pyrethrin II", "Pyridaben", "Quintozene", "Resmethrin", "Spinetoram", "Spinosad",
    "Spirodiclofen", "Spiromesifen", "Spirotetramat", "Spiroxamine", "Tebuconazole", "Tebufenozide",
    "Teflubenzuron", "Tetrachlorvinphos", "Tetramethrin", "Thiacloprid", "Thiamethoxam", "Thiophanate-methyl",
    "Trifloxystrobin", "Aflatoxin B1", "Aflatoxin B2", "Aflatoxin G1", "Aflatoxin G2", "Ochratoxin",
    "Zearalenone", "Captan", "Fenhexamid", "Chlordane", "TPP (IS)"
]

compoundDictonary = {}

for i, value in enumerate(values, start=1):
    compoundDictonary[i] = value


def scanPest(fileLocation): 
    print('SCANNING PEST EXCEL FILE')
    print('FILE LOCATION: ', fileLocation)
    
    wb = load_workbook(fileLocation, data_only=True) 
    ws = wb.active

    sampleNames = [] 
    jobNumbers = []
    
    column_letter = 'B'
    column_index = column_index_from_string(column_letter)
    starting_row = 76

    column_data = []
    temp = []

    sectionRows = []
    headerRows = []
    
    for row in ws.iter_rows(min_col=column_index, max_col=column_index, min_row=starting_row):
        for cell in row:
            
            if cell.value is not None:
                temp.append((cell.value, cell.row))
            
            if cell.value == 1: 
                sectionRows.append(cell.row)
                headerRows.append(cell.row-1) 
            
            if cell.value == 104: 
                
                sideCheck = ws.cell(row=cell.row, column=cell.column-1)
                if(sideCheck.value is None): 
                    column_data.append(temp)
                    temp = []
                
        
    lastRow = column_data[-1][1]
    
    print('Header Rows: {}'.format(headerRows)); 
    
    sampleNameLocations, sampleNames = determineSampleNumbers(ws,headerRows)

    print('**Sample Names')
    print(sampleNames)

    sampleData = {}
 
    for section in sampleNameLocations: 
        for sampleName, col in section: 
            sampleData[sampleName] = []
        
    print(sampleData)

    for i, section in enumerate(column_data):
        print('***Section: ', i)
        currentSectionLength = len(sampleNameLocations[i])
        print('***Total Columns: ', currentSectionLength)

        for testNum, testRow in section: 
            testName = compoundDictonary[testNum]
            print(f"**Test Name: {testName}, Test Number: {testNum}, Row: {testRow}")
    
            for j in range(currentSectionLength): 
                sampleName = sampleNameLocations[i][j][0]
                sampleCol = sampleNameLocations[i][j][1]
                sampleVal = ws.cell(row=testRow, column=sampleCol).value 
                
                if(sampleVal is None):  
                    sampleData[sampleName].append([testName, 'ND'])
                    print(f"SampleName: {sampleName}, sampleVal: {'ND'}")
                else: 
                    sampleData[sampleName].append([testName, float(sampleVal)])
                    print(f"SampleName: {sampleName}, sampleVal: {sampleVal}")
                    
        
    #bashTableOutput(sampleData, column_data)

    recoveryValues = determineRecoveryValues(ws, column_data[0]);

    jobNumbers = []
    for sampleName in sampleNames: 
        jobNumber = sampleName[:6]
        if(jobNumber not in jobNumbers):
            jobNumbers.append(jobNumber)
    
    
    return jobNumbers, sampleNames, sampleData, recoveryValues; 

    

def bashTableOutput(sampleData, column_data):
    print()
    print('Compound Name       |', end='') 
    placement = 4;
    for key in sampleData.keys(): 
        print(f' {key} |', end='') #11 or 12 
    
    print()
    print(272 * '-')

    for testNum, testRow in column_data[0]: 
        testName = compoundDictonary[testNum]
        wordLength = len(testName[:19])
        remainder = 19 - wordLength
        print(f"{testName[:19]:19s} |", end='')
        
        for i, key in enumerate(sampleData.keys(), start=1): 
            sampleValue = sampleData[key][testNum][1]
            if(sampleValue != 'ND'): 

                temp = '{:.4f}'.format(float(sampleValue))
                formattedValue = '{:.5g}'.format(float(temp))
                totalLength = len(formattedValue)
                
                if(len(key) == 8): 
                    remainder = 11 - (3 + totalLength) - 2
                else: 
                    remainder = 12 - (3 + totalLength) - 2
                
                print(3 * ' ' + formattedValue + (remainder * ' '), '|', end='')
            else: 
                
                if(len(key) == 8): 
                    remainder = 4
                else: 
                    remainder = 5; 
                
                print(4 * ' ' + 'ND' +  (remainder * ' ') +'|', end ='') 
            
        print()

def determineRecoveryColumn(ws): 
    headerRow = 74 
    
    for cell in ws[headerRow]: 
        if cell.value == 'CS5': 
            print('Recovery Column: ', cell.column)
            return cell.column

def determineRecoveryValues(ws, column_data): 
    recoveryColumn = determineRecoveryColumn(ws); 
    recoveryData = {}

    for testNum, testRow in column_data: 
        testName = compoundDictonary[testNum]
        recoveryData[testName] = ws.cell(row=testRow, column=recoveryColumn).value 
        
    return recoveryData
    
          
def determineSampleNumbers(ws, headerRows):

    sampleNames = []
    sampleNamesData = []
    
    for currentHeaderRow in headerRows:
        currentRowSampleNames = []
        for cell in ws[currentHeaderRow]:
            if cell.value == 'ng/g':
                row_above = currentHeaderRow - 1
                value_above = ws.cell(row=row_above, column=cell.column).value
                currentRowSampleNames.append([value_above, cell.column])
                
                if(value_above not in sampleNames): 
                    sampleNames.append(value_above)
                  
        sampleNamesData.append(currentRowSampleNames)
        
    return sampleNamesData, sampleNames 

    
                
    
    
    

    
    
    
    
    
